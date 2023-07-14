import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side
from openpyxl.utils import get_column_letter

folder_path = os.path.join(os.path.expanduser("~"), "Desktop", "OfficeCSV")

columns_to_read = ['Nome para exibição', 'Licenças', 'Nome UPN']

for file_name in os.listdir(folder_path):
    if file_name.endswith('.csv'):
        file_path = os.path.join(folder_path, file_name)

        df = pd.read_csv(file_path, usecols=columns_to_read)

        df = df.dropna(subset=['Licenças'])

        main_file_name = os.path.splitext(file_name)[0]

        excel_writer = pd.ExcelWriter(main_file_name + '.xlsx')
        df.to_excel(excel_writer, sheet_name='Dados', index=False)
        # excel_writer.save()
        excel_writer.close()

        book = Workbook()
        book = openpyxl.load_workbook(main_file_name + '.xlsx')
        sheet = book.active

        fill_color1 = "FFFFFF"  # Branco
        fill_color2 = "E1F0E1"  # Verde claro

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color1, end_color=fill_color1, fill_type="solid")
                cell.border = cell.border.copy(left=Side(), right=Side(), top=Side(), bottom=Side())

            fill_color1, fill_color2 = fill_color2, fill_color1

        max_col = df.shape[1]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=max_col + 1, max_col=sheet.max_column):
            for cell in row:
                cell.border = cell.border.copy(left=Side(), right=Side(), top=Side(), bottom=Side())

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        book.save(main_file_name + '.xlsx')
        book.close()
